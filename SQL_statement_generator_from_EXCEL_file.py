import sys
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime


##  A python script that analyzes an excel file and generates a SQL statement for existing DB with multiple tables.  
##  The script should be called from a CLI interface with three arguments.
##  1. `EXCEL` file path. 
##  2. Club id.
##  3. PK field title.

def main(excel_file_path, club_id, pk_title):
##  The main method of the script.

    try:
        workbook = openpyxl.load_workbook(excel_file_path)
    except:
        print(f"File path error!\n\
Could not open the file `{excel_file_path}`.\n\
Aborting script!")
        sys.exit()

    sheet = workbook.worksheets[0]
    pk_col_as_int = pk_validation_col_exist(sheet, pk_title)
    pk_validation_double_value_in_file(sheet, pk_col_as_int)

    row_values_dictonary = { 'club_id' : club_id }
    num_of_cols = sheet.max_column
    num_of_rows = sheet.max_row
    users_sql_statement = "INSERT INTO users (first_name, last_name, phone, email, joined_at, club_id) \nVALUES \n"
    memberships_sql_statement = "INSERT INTO memberships (user_id, start_date, end_date, membership_name) \nVALUES \n"

    for i in range (2, num_of_rows + 1):
        for j in range (1, num_of_cols + 1):
            column_title = sheet.cell(row = 1, column = j).value
            cell_value = sheet.cell(row = i, column = j).value

            if type(cell_value) == datetime:
                row_values_dictonary[column_title] = cell_value.date()
            else:
                if cell_value is None:
                    row_values_dictonary[column_title] = "NULL"
                else:    
                    row_values_dictonary[column_title] = cell_value

        users_sql_statement = add_row_to_users_statement(users_sql_statement, row_values_dictonary)
        memberships_sql_statement = add_row_to_memberships_statement(memberships_sql_statement, row_values_dictonary, pk_title)

    users_sql_statement = correct_sql_statement(users_sql_statement)
    memberships_sql_statement = correct_sql_statement(memberships_sql_statement)

    print(f"{users_sql_statement}\n\n{memberships_sql_statement}")

def pk_validation_col_exist(sheet, pk_title):
## The method gets sheet and PK title.
## If the PK column doesn't exist in the sheet, abort the script.
    pk_col_index = get_title_col_index(sheet, pk_title)

    if pk_col_index == -1:
        print(f"No `{pk_title}` column in the file!\nAborting script")
        sys.exit()

    return pk_col_index

def get_title_col_index(sheet, title):
## The method gets sheet and title.
## Returns the index of the title's col, if no such a title, returns -1.

    num_of_cols = sheet.max_column
    for i in range (1, num_of_cols + 1):
        if sheet.cell(row = 1, column = i).value == title:
            return i
    return -1

def pk_validation_double_value_in_file(sheet, pk_col_index):
##  The method gets sheet and PK column index.
##  If any PK value exist double or more times, abort the script.

    pk_dictonary = {}

    pk_col_as_letter = get_column_letter(pk_col_index)
    for pk in sheet[pk_col_as_letter]:
        if pk_dictonary.__contains__(pk.value) is False:
            pk_dictonary[pk.value] = pk.value
        else:
            print(f"The PK value - `{pk.value}`, exists DOUBLE or more times in the file.\nAborting script")
            sys.exit()

def add_row_to_users_statement(users_sql_statement, row_values_dictonary):
## The method gets 'users' SQL statement and row values.
## Returns a combined SQL statement of the statement and the row values.

    users_row = (f"      (\
'{row_values_dictonary['first_name']}', '{row_values_dictonary['last_name']}',\
 '{row_values_dictonary['phone']}', '{row_values_dictonary['email']}',\
 '{row_values_dictonary['membershp_start_date']}', {row_values_dictonary['club_id']}),\n")

    users_sql_statement += users_row
    return users_sql_statement

def add_row_to_memberships_statement(memberships_sql_statement, row_values_dictonary, pk_title):
## The method gets 'memberships' SQL statement, row values and PK title.
## Returns a combined SQL statement of the statement and the row values using the PK title.

    memberships_row = (f"      (\
SELECT id FROM users WHERE {pk_title}={row_values_dictonary[pk_title]};, '{row_values_dictonary['membershp_start_date']}',\
 '{row_values_dictonary['membership_end_date']}', '{row_values_dictonary['membership_name']}'),\n")
 
    memberships_sql_statement += memberships_row
    return memberships_sql_statement

def correct_sql_statement(sql_statement):
##  The method gets a string representing invalid SQL statement - extra ',' and '\n' also missing ';'.
##  Returns a string representing a valid SQL statement.

    sql_statement = sql_statement[0:-2]
    sql_statement += ';'
    return sql_statement

if __name__== "__main__":
    main(sys.argv[1], sys.argv[2], sys.argv[3])